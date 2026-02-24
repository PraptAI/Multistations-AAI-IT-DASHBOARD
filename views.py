
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.contrib import messages
from django.http import HttpResponse


from .forms import (
    ITCoordinatorForm, ContractStaffForm, AssetForm,
    AssetTypeForm, DepartmentForm, DesignationForm, StationForm,
    ContractForm, EmployeeDirectoryForm, LocationForm, AssetIssuanceForm
)
from .models import (
    ITCoordinator, ContractStaff, Station, UserProfile, Asset, Location,
    AssetType, Department, Designation, Contract, EmployeeDirectory, AssetIssuance
)


#  DASHBOARD 
@login_required
def dashboard(request):
    user_station_name = None
    if hasattr(request.user, "userprofile"):
        station = getattr(request.user.userprofile, "station", None)
        if station:
            user_station_name = station.name

    return render(request, "accounts/dashboard.html", {
        "user_station_name": user_station_name
    })


# LOGIN 
from django.contrib.auth.views import LoginView
from django.contrib import messages
from .models import Station

class CustomLoginView(LoginView):
    template_name = 'accounts/login.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stations'] = Station.objects.all()
        return context

    def form_invalid(self, form):
        messages.error(self.request, "Invalid username or password")
        return super().form_invalid(form)


# SIGNUP 
import re
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import login
from django.shortcuts import render, redirect
from .models import Station, UserProfile


def signup(request):
    stations = Station.objects.filter(userprofile__isnull=True)

    if request.method == "POST":
        username = request.POST['username']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        station_id = request.POST['station']
        is_rhqwr = request.POST.get('is_rhqwr') == 'on'

        if password1 != password2:
            messages.error(request, "Passwords do not match")
            return redirect('signup')

        #  PASSWORD RULES

        if len(password1) < 8:
            messages.error(request, "Password must be at least 8 characters long")
            return redirect('signup')

        if not re.search(r"[A-Z]", password1):
            messages.error(request, "Password must contain at least one capital letter")
            return redirect('signup')

        if not re.search(r"[a-z]", password1):
            messages.error(request, "Password must contain at least one small letter")
            return redirect('signup')

        if not re.search(r"[0-9]", password1):
            messages.error(request, "Password must contain at least one number")
            return redirect('signup')

        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password1):
            messages.error(request, "Password must contain at least one special character")
            return redirect('signup')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return redirect('signup')

        if UserProfile.objects.filter(station_id=station_id).exists():
            messages.error(request, "This station already has a login account")
            return redirect('signup')

        user = User.objects.create_user(
            username=username,
            password=password1
        )

        UserProfile.objects.create(
            user=user,
            station_id=station_id,
            is_rhqwr=is_rhqwr
        )

        login(request, user)
        return redirect('dashboard')

    return render(request, 'accounts/signup.html', {'stations': stations})


#  IT COORDINATOR 
@login_required
def list_it_coordinators(request):
    user = request.user
    user_station = getattr(user.userprofile, 'station', None)

    
    is_rhqwr = False
    if user_station:
        station_normalized = user_station.name.replace(" ", "").replace("-", "").upper()
        if station_normalized in ["RHQWR", "RHQWESTERNREGION"]:
            is_rhqwr = True

    coordinators = ITCoordinator.objects.all()

    return render(request, "accounts/list_it_coordinators.html", {
        "coordinators": coordinators,
        "is_rhqwr": is_rhqwr or user.is_superuser,
        "user_station": user_station,
    })


# ADD IT COORDINATOR
@login_required
def add_it_coordinator(request):
    if request.method == "POST":
        form = ITCoordinatorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("dashboard")
    else:
        form = ITCoordinatorForm()
        if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = form.fields['station'].queryset.filter(
                pk=request.user.userprofile.station.pk
            )
    return render(request, "accounts/add_it_coordinator.html", {"form": form})

# EDIT IT COORDINATOR
@login_required
def edit_it_coordinator(request, pk):
    coordinator = get_object_or_404(ITCoordinator, pk=pk)
    if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
        if coordinator.station != request.user.userprofile.station:
            return redirect("list_it_coordinators")

    if request.method == "POST":
        form = ITCoordinatorForm(request.POST, instance=coordinator)
        if form.is_valid():
            form.save()
            return redirect("list_it_coordinators")
    else:
        form = ITCoordinatorForm(instance=coordinator)
        if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = form.fields['station'].queryset.filter(
                pk=request.user.userprofile.station.pk
            )
    return render(request, "accounts/edit_it_coordinator.html", {"form": form})


#  CONTRACT STAFF 

def list_contract_staff(request):
    user_station = getattr(request.user.userprofile, 'station', None)
    if request.user.is_superuser or (user_station and user_station.name == "RHQ WR"):
        staff = ContractStaff.objects.all()
    else:
        staff = ContractStaff.objects.filter(station=user_station)
    return render(request, "accounts/list_contract_staff.html", {"staff": staff})

# ADD CONTRACT STAFF

@login_required
def add_contract_staff(request):
    if request.method == "POST":
        form = ContractStaffForm(request.POST)
        if form.is_valid():
            contract_staff = form.save(commit=False)
            if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
                contract_staff.station = request.user.userprofile.station
            contract_staff.save()
            return redirect("list_contract_staff")
    else:
        form = ContractStaffForm()
        if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = form.fields['station'].queryset.filter(
                pk=request.user.userprofile.station.pk
            )
    return render(request, "accounts/add_contract_staff.html", {"form": form})

# EDIT CONTRACT STAFF

@login_required
def edit_contract_staff(request, pk):
    staff = get_object_or_404(ContractStaff, pk=pk)
    if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
        if staff.station != request.user.userprofile.station:
            return redirect("list_contract_staff")

    if request.method == "POST":
        form = ContractStaffForm(request.POST, instance=staff)
        if form.is_valid():
            form.save()
            return redirect("list_contract_staff")
    else:
        form = ContractStaffForm(instance=staff)
        if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = form.fields['station'].queryset.filter(
                pk=request.user.userprofile.station.pk
            )
    return render(request, "accounts/edit_contract_staff.html", {"form": form})


#  ASSET MANAGEMENT 

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Asset, Station  

@login_required
def list_assets(request):
    user = request.user
    user_profile = getattr(user, "userprofile", None)

    assets = Asset.objects.all()
    stations = None

    #  STATION RESTRICTION

    if not user.is_superuser and not user_profile.is_rhqwr:
        assets = assets.filter(station=user_profile.station)

    #  RHQ-WR 
    if user.is_superuser or user_profile.is_rhqwr:
        stations = Station.objects.all()
        station_id = request.GET.get("station")

        if station_id:
            assets = assets.filter(station_id=station_id)

    context = {
        "assets": assets,
        "stations": stations,
    }
    return render(request, "accounts/list_assets.html", context)




#  ASSET TYPE CRUD 

def add_asset_type(request):
    if request.method == "POST":
        form = AssetTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('add_asset_type')
    else:
        form = AssetTypeForm()

    asset_types = AssetType.objects.all()
    return render(request, "accounts/add_asset_type.html", {
        "form": form,
        "asset_types": asset_types
    })

# EDIT ASSET TYPE

@login_required
def edit_asset_type(request, pk):
    asset_type = get_object_or_404(AssetType, pk=pk)

    if request.method == "POST":
        form = AssetTypeForm(request.POST, instance=asset_type)
        if form.is_valid():
            form.save()
            return redirect('add_asset_type')
    else:
        form = AssetTypeForm(instance=asset_type)

    asset_types = AssetType.objects.all()
    return render(request, "accounts/add_asset_type.html", {
        "form": form,
        "asset_types": asset_types
    })


#  DEPARTMENT , DESIGNATION AND STATION

def add_department(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('add_department')
    else:
        form = DepartmentForm()
    departments = Department.objects.all()
    return render(request, "accounts/add_department.html", {'form': form, 'departments': departments})


def add_designation(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('add_designation')
    else:
        form = DesignationForm()
    designations = Designation.objects.all()
    return render(request, "accounts/add_designation.html", {'form': form, 'designations': designations})


def add_station(request):
    if request.method == 'POST':
        form = StationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('add_station')
    else:
        form = StationForm()
    stations = Station.objects.all()
    return render(request, "accounts/add_station.html", {'form': form, 'stations': stations})

# LOGOUT 

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

# EDIT STATION 

@login_required
def edit_station(request, station_id):
    station = get_object_or_404(Station, id=station_id)

    if request.method == "POST":
        form = StationForm(request.POST, instance=station)
        if form.is_valid():
            form.save()
            messages.success(request, "Station updated successfully!")
            return redirect('add_station')
    else:
        form = StationForm(instance=station)

    stations = Station.objects.all()
    return render(request, 'accounts/add_station.html', {
        'form': form,
        'stations': stations,
    })

# EDIT DEPARTMENT 
@login_required
def edit_department(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ Department updated successfully!")
            return redirect("add_department")
    else:
        form = DepartmentForm(instance=department)

    departments = Department.objects.all().order_by("name")
    return render(request, "accounts/add_department.html", {
        "form": form,
        "departments": departments,
    })


@login_required
def edit_designation(request, pk):
    designation = get_object_or_404(Designation, pk=pk)

    if request.method == "POST":
        form = DesignationForm(request.POST, instance=designation)
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ Designation updated successfully!")
            return redirect("add_designation")
    else:
        form = DesignationForm(instance=designation)

    designations = Designation.objects.all().order_by("name")
    return render(request, "accounts/add_designation.html", {
        "form": form,
        "designations": designations,
    })


#  ADD CONTRACT
@login_required
def add_contract(request):
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            contract = form.save(commit=False)
            if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
                contract.station = request.user.userprofile.station
            contract.save()
            messages.success(request, "✅ Contract added successfully.")
            return redirect('list_contracts')
        else:
            messages.error(request, "⚠️ Error: Please check the form.")
    else:
        form = ContractForm()
        if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = Station.objects.filter(
                pk=request.user.userprofile.station.pk
            )

    return render(request, 'accounts/add_contract.html', {'form': form})


@login_required
def list_contracts(request):
    user_station = getattr(request.user.userprofile, "station", None)

    if request.user.is_superuser or request.user.userprofile.is_rhqwr:
        contracts = Contract.objects.all()
    else:
        contracts = Contract.objects.filter(station=user_station)

    contracts = contracts.order_by('-contract_start_date')
    return render(request, "accounts/list_contracts.html", {'contracts': contracts})


@login_required
def edit_contract(request, contract_id):
    contract = get_object_or_404(Contract, id=contract_id)

    if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
        if contract.station != request.user.userprofile.station:
            messages.error(request, "❌ You cannot edit contracts from other stations.")
            return redirect("list_contracts")

    if request.method == "POST":
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            contract_obj = form.save(commit=False)
            if not request.user.is_superuser and not request.user.userprofile.is_rhqwr:
                contract_obj.station = request.user.userprofile.station
            contract_obj.save()
            messages.success(request, "✅ Contract updated successfully.")
            return redirect("list_contracts")
    else:
        form = ContractForm(instance=contract)

    return render(request, 'accounts/edit_contract.html', {'form': form})


#  Employee Directory 
@login_required
def add_employee_directory(request):
    if request.method == 'POST':
        form = EmployeeDirectoryForm(request.POST)
        if form.is_valid():
            emp = form.save(commit=False)
            if not request.user.is_superuser and getattr(request.user, 'userprofile', None) and not request.user.userprofile.is_rhqwr:
                emp.station = request.user.userprofile.station
            emp.save()
            messages.success(request, "Employee added to directory.")
            return redirect('list_employee_directory')
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = EmployeeDirectoryForm()
        if not request.user.is_superuser and getattr(request.user, 'userprofile', None) and not request.user.userprofile.is_rhqwr:
            form.fields['station'].queryset = form.fields['station'].queryset.filter(
                pk=request.user.userprofile.station.pk
            )
    return render(request, 'accounts/add_employee_directory.html', {'form': form})


@login_required
def list_employee_directory(request):
    if request.user.is_superuser or getattr(request.user.userprofile, 'is_rhqwr', False):
        employees = EmployeeDirectory.objects.all().select_related('department', 'designation', 'station')
    else:
        station = getattr(request.user.userprofile, 'station', None)
        employees = EmployeeDirectory.objects.filter(station=station).select_related('department', 'designation', 'station')
    return render(request, 'accounts/list_employee_directory.html', {'employees': employees})


@login_required
def edit_employee_directory(request, pk):
    employee = get_object_or_404(EmployeeDirectory, pk=pk)

    if request.method == 'POST':
        form = EmployeeDirectoryForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Employee details updated successfully.")
            return redirect('list_employee_directory')
    else:
        form = EmployeeDirectoryForm(instance=employee)

    return render(request, 'accounts/edit_employee_directory.html', {'form': form})


#  Location 
def add_location(request):
    user_station = request.user.userprofile.station
    if request.method == "POST":
        form = LocationForm(request.POST)
        if form.is_valid():
            location = form.save(commit=False)
            location.station = user_station
            location.save()
            return redirect('add_location')
    else:
        form = LocationForm()

    locations = Location.objects.filter(station=user_station)
    return render(request, "accounts/add_location.html", {
        "form": form,
        "locations": locations
    })


def edit_location(request, id):
    user_station = request.user.userprofile.station
    location = get_object_or_404(Location, id=id)

    if location.station != user_station:
        return HttpResponse("Access Denied")

    if request.method == "POST":
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            return redirect('add_location')
    else:
        form = LocationForm(instance=location)

    locations = Location.objects.filter(station=user_station)
    return render(request, "accounts/add_location.html", {
        "form": form,
        "locations": locations
    })


#  ASSETS 

def add_asset(request):
    if request.method == "POST":
        form = AssetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Asset added successfully!")
            return redirect('list_assets')
    else:
        form = AssetForm()

    return render(request, 'accounts/add_asset.html', {'form': form})


@login_required
def edit_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    user_profile = getattr(request.user, "userprofile", None)
    user_station = getattr(user_profile, "station", None)

    # Restrict non-RHQ users from editing other stations
    if not request.user.is_superuser and not user_profile.is_rhqwr:
        if asset.station != user_station:
            return redirect("list_assets")

    if request.method == "POST":
        
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            asset_obj = form.save(commit=False)

            location_name = form.cleaned_data.get("location_name")
            station = form.cleaned_data.get("station")

            
            if not request.user.is_superuser and not user_profile.is_rhqwr:
                station = user_station

            asset_obj.station = station

            
            if location_name:
                location_obj, _ = Location.objects.get_or_create(
                    name=location_name,
                    station=station
                )
                asset_obj.location = location_obj

            asset_obj.save()
            return redirect("list_assets")

    else:
        initial = {
            "location_name": asset.location.name if asset.location else ""
        }
        form = AssetForm(instance=asset, initial=initial)

    return render(request, "accounts/edit_asset.html", {"form": form})


#  ASSET ISSUANCE 
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Asset, AssetIssuance
from .forms import AssetIssuanceForm

@login_required
def new_issuance(request):
    asset_details = None
    asset_already_issued = False
    user_station = getattr(request.user.userprofile, "station", None)
    is_rhqwr = getattr(request.user.userprofile, "is_rhqwr", False)

    serial_number = request.POST.get("serial_number") or request.GET.get("serial_number")

    # Fetch asset details if serial number provided
    if serial_number:
        try:
            asset = Asset.objects.get(serial_number=serial_number)

            # Restrict station for non-superuser/non-RHQ
            if not request.user.is_superuser and not is_rhqwr and asset.station != user_station:
                messages.error(request, "❌ You cannot issue assets from other stations.")
            else:
                asset_details = asset

                # Check if asset is currently issued
                if AssetIssuance.objects.filter(serial_number=serial_number, issued_status="user").exists():
                    asset_already_issued = True
                    messages.error(request, f"⚠️ Asset '{asset.asset_type} - {serial_number}' is already issued!")

        except Asset.DoesNotExist:
            messages.error(request, f"❌ No asset found with serial number {serial_number}.")

    # Initialize form with asset details
    initial_data = {}
    if asset_details:
        initial_data = {
            "make": asset_details.make,
            "model": asset_details.model,
            "asset_type": asset_details.asset_type,
            "location": asset_details.location,
            "aai_asset_code": asset_details.aai_asset_code,
            "station": asset_details.station,
            "serial_number": asset_details.serial_number,
        }

    form = AssetIssuanceForm(request.POST or None, initial=initial_data, user=request.user)

    # Handle form submission
    if request.method == "POST" and form.is_valid() and asset_details and not asset_already_issued:
        #  Create a new issuance row instead of overwriting
        AssetIssuance.objects.create(
    serial_number=asset_details.serial_number,
    station=asset_details.station,
    location=asset_details.location,
    asset_type=asset_details.asset_type,
    make=asset_details.make,
    model=asset_details.model,
    aai_asset_code=asset_details.aai_asset_code,

    user_name=form.cleaned_data.get("user_name"),
    user_department=form.cleaned_data.get("user_department"),
    user_designation=form.cleaned_data.get("user_designation"),
    user_extension=form.cleaned_data.get("user_extension"),

    issued_status="user",
    is_active=True,
    date_of_issue=form.cleaned_data.get("date_of_issue"),
)



        messages.success(
            request,
            f" Asset '{asset_details.asset_type} - {asset_details.serial_number}' issued successfully!"
        )
        return redirect("issued_asset_list")

    context = {
        "form": form,
        "asset_details": asset_details,
        "asset_already_issued": asset_already_issued,
    }
    return render(request, "accounts/new_issuance.html", context)





from django.contrib.auth.decorators import login_required
from .models import AssetIssuance, Station

@login_required
def issued_asset_list(request):
    user_profile = getattr(request.user, "userprofile", None)

    
    issued_assets = AssetIssuance.objects.filter(issued_status="user")

    if not request.user.is_superuser and not getattr(user_profile, "is_rhqwr", False):
        if user_profile and user_profile.station:
            issued_assets = issued_assets.filter(station=user_profile.station)
        else:
            issued_assets = issued_assets.none()

    
    station_id = request.GET.get("station")
    if station_id and (request.user.is_superuser or getattr(user_profile, "is_rhqwr", False)):
        issued_assets = issued_assets.filter(station_id=station_id)

    issued_assets = issued_assets.order_by("-date_of_issue")

    context = {
        "issued_assets": issued_assets,
        "stations": Station.objects.all(),   
        "total_issued_assets": issued_assets.count(),
    }

    return render(request, "accounts/issued_asset_list.html", context)




def asset_details_json(request, serial):
    try:
        asset = Asset.objects.get(serial_number=serial)

        issued = AssetIssuance.objects.filter(
            serial_number=serial,
            issued_status="user"
        ).exists()

        return JsonResponse({
            "exists": True,
            "make": asset.make or "",
            "model": asset.model or "",
            "station": asset.station.name if asset.station else "",
            "station_id": asset.station.id if asset.station else None,
            "location": asset.location.name if asset.location else "",
            "location_id": asset.location.id if asset.location else None,
            "asset_type_id": asset.asset_type.id if asset.asset_type else None,
            "aai_asset_code": asset.aai_asset_code or "",
            "issued_status": "user" if issued else "stock",
        })

    except Asset.DoesNotExist:
        return JsonResponse({"exists": False})



from django.contrib.auth.decorators import login_required





@login_required
def edit_issued_asset(request, pk):
    issuance = get_object_or_404(AssetIssuance, pk=pk)
    user_profile = getattr(request.user, "userprofile", None)

    
    if not request.user.is_superuser and not getattr(user_profile, "is_rhqwr", False):
        if not user_profile or issuance.station != user_profile.station:
            messages.error(request, "You cannot edit assets from other stations.")
            return redirect("issued_asset_list")

    if request.method == "POST":
        form = AssetIssuanceForm(
            request.POST,
            instance=issuance,
            user=request.user
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Issued asset updated successfully!")
            return redirect("issued_asset_list")
    else:
        form = AssetIssuanceForm(
            instance=issuance,
            user=request.user
        )

    return render(
        request,
        "accounts/edit_issued_asset.html",
        {"form": form}
    )


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils.timezone import now
from .models import AssetIssuance

@login_required
def return_asset(request, pk):
    # Fetch the issuance row that is currently issued
    issuance = get_object_or_404(
        AssetIssuance,
        pk=pk,
        issued_status="user"  # currently issued
    )

    user_profile = getattr(request.user, "userprofile", None)

    #  Restrict by station for non-superuser/non-RHQ
    if not request.user.is_superuser and not getattr(user_profile, "is_rhqwr", False):
        if issuance.station != user_profile.station:
            messages.error(request, " You cannot return assets from another station.")
            return redirect("issued_asset_list")

    if request.method == "POST":
        # Update the issuance row to mark it returned
        issuance.issued_status = "stock"  
        issuance.date_of_return = request.POST.get("return_date") or now().date()
        issuance.return_remarks = request.POST.get("return_remarks")
        issuance.is_active = False  

        issuance.save()

        messages.success(request, f"✅ Asset '{issuance.serial_number}' returned successfully!")
        return redirect("issued_asset_list")

    return render(
        request,
        "accounts/return_asset.html",
        {"issuance": issuance}
    )



from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import AssetIssuance, Station

@login_required
def asset_history(request):
    user_profile = getattr(request.user, "userprofile", None)


    history = AssetIssuance.objects.all()

    #  Restrict for non-superuser & non-RHQWR users
    if not request.user.is_superuser and not getattr(user_profile, "is_rhqwr", False):
        if user_profile and user_profile.station:
            history = history.filter(station=user_profile.station)
        else:
            history = history.none()

    #  Station filter  RHQWR)
    station_id = request.GET.get("station")
    if station_id and (request.user.is_superuser or getattr(user_profile, "is_rhqwr", False)):
        history = history.filter(station_id=station_id)

    history = history.order_by('-date_of_issue')

    context = {
        "history": history,
        "stations": Station.objects.all(),    
        "total_assets_history": history.count(),
    }

    return render(request, "accounts/asset_history.html", context)




@login_required
def asset_detail_history(request, serial_number):
    history = AssetIssuance.objects.filter(
        serial_number=serial_number
    ).order_by('-date_of_issue')

    return render(request, "accounts/asset_detail_history.html", {
        "serial_number": serial_number,
        "history": history
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from .models import Ticket
from .forms import TicketForm
from accounts.models import EmployeeDirectory
from django.contrib.auth.models import User

#  CREATE TICKET

@login_required
def create_ticket(request):
    user = request.user
    station = user.userprofile.station

    #  Engineer queryset (Station-wise, RHQ sees only RHQ engineers)
    engineers = Engineer.objects.filter(
        station=station,
        is_active=True
    )

    if request.method == 'POST':
        form = TicketForm(request.POST, user=user)
        form.fields['assigned_engineer'].queryset = engineers

        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = user

            #  Station-wise Ticket Number Generation
            last_ticket = Ticket.objects.filter(
                ticket_number__startswith=f"{station}-"
            ).order_by('-id').first()

            if last_ticket and last_ticket.ticket_number:
                try:
                    last_no = int(last_ticket.ticket_number.split('-')[1])
                except (IndexError, ValueError):
                    last_no = 0
            else:
                last_no = 0

            new_no = last_no + 1
            ticket.ticket_number = f"{station}-{new_no:03d}"

            ticket.save()

            messages.success(
                request,
                f"Ticket {ticket.ticket_number} created successfully."
            )
            return redirect('ticket_list')

    else:
        form = TicketForm(user=user)
        form.fields['assigned_engineer'].queryset = engineers

    return render(
        request,
        'accounts/create_ticket.html',
        {'form': form}
    )




#  LIST TICKETS (OPEN / CLOSED)

from django.utils import timezone
from datetime import date

from django.utils import timezone
from datetime import datetime, time

@login_required
def ticket_list(request):
    status = request.GET.get('status', 'open')
    station = request.user.userprofile.station

    today = timezone.localdate()

    start_datetime = timezone.make_aware(
        datetime.combine(today, time.min)
    )
    end_datetime = timezone.make_aware(
        datetime.combine(today, time.max)
    )

    tickets = Ticket.objects.filter(
        requester__station=station,
        call_time__range=(start_datetime, end_datetime)
    )

    if status == 'closed':
        tickets = tickets.filter(closed_at__isnull=False)
    else:
        tickets = tickets.filter(closed_at__isnull=True)

    tickets = tickets.order_by('-call_time')

    return render(
        request,
        'accounts/ticket_list.html',
        {
            'tickets': tickets,
            'active_status': status
        }
    )




#  CLOSE TICKET

from datetime import datetime
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone

@login_required
def close_ticket(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)

    if request.method == "POST":
        resolution = request.POST.get("resolution", "").strip()
        closed_at_str = request.POST.get("closed_at")

        #  Resolution validation
        if not resolution:
            messages.error(request, "Resolution is required.")
            return redirect("close_ticket", pk=pk)

        #  Closing time validation
        try:
            if closed_at_str:
                naive_dt = datetime.strptime(closed_at_str, "%Y-%m-%dT%H:%M")
                closed_at = timezone.make_aware(naive_dt)
            else:
                closed_at = timezone.now()
        except ValueError:
            messages.error(request, "Invalid closing time format.")
            return redirect("close_ticket", pk=pk)

        #  Logical validation
        if ticket.call_time and closed_at < ticket.call_time:
            messages.error(
                request,
                "Closing time cannot be earlier than Call Time."
            )
            return redirect("close_ticket", pk=pk)

        #  Save ticket
        ticket.closed_at = closed_at
        ticket.resolution = resolution
        ticket.closed_by = request.user
        ticket.save()

        messages.success(request, "Ticket closed successfully.")
        return redirect("ticket_list")

    return render(request, "accounts/close_ticket.html", {
        "ticket": ticket
    })


#  Employee Details

@login_required
def get_employee_details(request):
    emp_id = request.GET.get('employee_id')
    if not emp_id:
        return JsonResponse({'error': 'No employee selected'}, status=400)

    try:
        employee = EmployeeDirectory.objects.get(id=emp_id)
        data = {
            'department': employee.department.name if employee.department else '',
            'station': employee.station.name if employee.station else '',
            'designation': employee.designation.name if employee.designation else '',
            'phone': employee.mobile or employee.direct_line_phone or '',
            'email': employee.official_email or '',
        }
        return JsonResponse(data)
    except EmployeeDirectory.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)





from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import F
from openpyxl import Workbook
import math
from .models import Ticket, Engineer
from django.contrib.auth.decorators import login_required
from datetime import datetime, time, date


@login_required
def ticket_report(request):
    user = request.user

    # Base queryset (station filter)
    tickets = Ticket.objects.filter(
        requester__station=user.userprofile.station
    )

   
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    status = request.GET.get("status")
    engineer_id = request.GET.get("engineer")

    
    if start_date:
        start = timezone.make_aware(
            datetime.combine(
                datetime.strptime(start_date, "%Y-%m-%d"),
                time.min
            )
        )
        tickets = tickets.filter(created_at__gte=start)

    if end_date:
        end = timezone.make_aware(
            datetime.combine(
                datetime.strptime(end_date, "%Y-%m-%d"),
                time.max
            )
        )
        tickets = tickets.filter(created_at__lte=end)

    # Status filter
    if status == "open":
        tickets = tickets.filter(closed_at__isnull=True)
    elif status == "closed":
        tickets = tickets.filter(closed_at__isnull=False)

    # Engineer filter
    if engineer_id:
        tickets = tickets.filter(assigned_engineer_id=engineer_id)

    # Engineers for dropdown
    if user.userprofile.station == "RHQ":
        engineers = Engineer.objects.filter(station="RHQ")
    else:
        engineers = Engineer.objects.filter(
            station=user.userprofile.station
        )

    tickets = tickets.order_by("-created_at")

    return render(request, "accounts/ticket_report.html", {
        "tickets": tickets,
        "engineers": engineers,
        "selected_engineer": engineer_id or "",
        "active_status": status or "",
        "start_date": start_date or "",
        "end_date": end_date or "",
    })




@login_required
def export_ticket_report(request):
    user = request.user

    tickets = Ticket.objects.filter(
        requester__station=user.userprofile.station
    )

    # GET params
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    status = request.GET.get("status")
    engineer_id = request.GET.get("engineer")

    from datetime import datetime, time


    if start_date:
        start = timezone.make_aware(
            datetime.combine(
                datetime.strptime(start_date, "%Y-%m-%d"),
                time.min
            )
        )
        tickets = tickets.filter(created_at__gte=start)

    if end_date:
        end = timezone.make_aware(
            datetime.combine(
                datetime.strptime(end_date, "%Y-%m-%d"),
                time.max
            )
        )
        tickets = tickets.filter(created_at__lte=end)

    # Status filter
    if status == "open":
        tickets = tickets.filter(closed_at__isnull=True)
    elif status == "closed":
        tickets = tickets.filter(closed_at__isnull=False)

    # Engineer filter
    if engineer_id:
        tickets = tickets.filter(assigned_engineer_id=engineer_id)

    # Excel creation
    wb = Workbook()
    ws = wb.active
    ws.title = "Ticket Report"

    ws.append([
        "Ticket ID", "Subject", "Requester", "Engineer",
        "Priority", "Status", "Created At", "Closed At", "Resolution Time"
    ])

    for t in tickets:
        created_local = timezone.localtime(t.created_at)
        closed_local = timezone.localtime(t.closed_at) if t.closed_at else None

        if closed_local:
            delta = closed_local - created_local
            total_minutes = math.ceil(delta.total_seconds() / 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            resolution_str = f"{hours}h {minutes}m"
        else:
            resolution_str = "-"

        ws.append([
            t.id,
            t.subject,
            t.requester.name,
            t.assigned_engineer.name if t.assigned_engineer else "-",
            t.get_priority_display(),
            "Closed" if t.closed_at else "Open",
            created_local.strftime("%d-%m-%Y %H:%M"),
            closed_local.strftime("%d-%m-%Y %H:%M") if closed_local else "-",
            resolution_str
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    today_str = date.today().strftime("%d-%m-%Y")
    response["Content-Disposition"] = f"attachment; filename=Ticket_Report_{today_str}.xlsx"

    wb.save(response)
    return response


from django.shortcuts import render
from django.db.models import F
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from django.utils import timezone
from datetime import datetime, time, date
import math

from .models import Ticket, Engineer


@login_required
def engineer_report(request):
    user = request.user

    # Engineers by station
    if user.userprofile.station == 'RHQ':
        engineers = Engineer.objects.filter(station='RHQ')
    else:
        engineers = Engineer.objects.filter(station=user.userprofile.station)

    selected_engineer = request.GET.get('engineer')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if selected_engineer:
        engineers = engineers.filter(id=selected_engineer)

    report_data = []

    for eng in engineers:
        tickets = Ticket.objects.filter(assigned_engineer=eng)

       
        if start_date:
            start = timezone.make_aware(
                datetime.combine(
                    datetime.strptime(start_date, "%Y-%m-%d"),
                    time.min
                )
            )
            tickets = tickets.filter(created_at__gte=start)

        if end_date:
            end = timezone.make_aware(
                datetime.combine(
                    datetime.strptime(end_date, "%Y-%m-%d"),
                    time.max
                )
            )
            tickets = tickets.filter(created_at__lte=end)

        total_calls = tickets.count()
        open_calls = tickets.filter(closed_at__isnull=True).count()
        closed_calls = tickets.filter(
            closed_at__isnull=False,
            call_time__isnull=False,
            closed_at__gte=F('call_time')
        )

        # Avg Resolution Time
        if closed_calls.exists():
            total_minutes_list = []
            for t in closed_calls:
                call_local = timezone.localtime(t.call_time)
                closed_local = timezone.localtime(t.closed_at)
                delta = closed_local - call_local
                total_minutes = math.ceil(delta.total_seconds() / 60)
                total_minutes_list.append(total_minutes)

            avg_total_minutes = sum(total_minutes_list) / len(total_minutes_list)
            avg_hours = int(avg_total_minutes) // 60
            avg_minutes = int(avg_total_minutes) % 60
            avg_resolution = f"{avg_hours}h {avg_minutes}m"
        else:
            avg_resolution = "—"

        report_data.append({
            'engineer': eng,
            'total_calls': total_calls,
            'open_calls': open_calls,
            'closed_calls': closed_calls.count(),
            'avg_resolution': avg_resolution,
        })

    return render(request, 'accounts/engineer_report.html', {
        'report_data': report_data,
        'engineers': engineers,
        'selected_engineer': selected_engineer or '',
        'start_date': start_date or '',
        'end_date': end_date or '',
    })





@login_required
def export_engineer_report(request):
    user = request.user

    if user.userprofile.station == 'RHQ':
        engineers = Engineer.objects.filter(station='RHQ')
    else:
        engineers = Engineer.objects.filter(station=user.userprofile.station)

    selected_engineer = request.GET.get('engineer')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if selected_engineer:
        engineers = engineers.filter(id=selected_engineer)

    wb = Workbook()
    ws = wb.active
    ws.title = "Engineer Report"

    ws.append([
        "Engineer",
        "Total Calls",
        "Open Calls",
        "Closed Calls",
        "Avg Resolution Time"
    ])

    for eng in engineers:
        tickets = Ticket.objects.filter(assigned_engineer=eng)

       
        if start_date:
            start = timezone.make_aware(
                datetime.combine(
                    datetime.strptime(start_date, "%Y-%m-%d"),
                    time.min
                )
            )
            tickets = tickets.filter(created_at__gte=start)

        if end_date:
            end = timezone.make_aware(
                datetime.combine(
                    datetime.strptime(end_date, "%Y-%m-%d"),
                    time.max
                )
            )
            tickets = tickets.filter(created_at__lte=end)

        total_calls = tickets.count()
        open_calls = tickets.filter(closed_at__isnull=True).count()
        closed_calls = tickets.filter(
            closed_at__isnull=False,
            call_time__isnull=False,
            closed_at__gte=F('call_time')
        )

        if closed_calls.exists():
            total_minutes_list = []
            for t in closed_calls:
                call_local = timezone.localtime(t.call_time)
                closed_local = timezone.localtime(t.closed_at)
                delta = closed_local - call_local
                total_minutes = math.ceil(delta.total_seconds() / 60)
                total_minutes_list.append(total_minutes)

            avg_total_minutes = sum(total_minutes_list) / len(total_minutes_list)
            avg_hours = int(avg_total_minutes) // 60
            avg_minutes = int(avg_total_minutes) % 60
            avg_resolution = f"{avg_hours}h {avg_minutes}m"
        else:
            avg_resolution = "—"

        ws.append([
            eng.name,
            total_calls,
            open_calls,
            closed_calls.count(),
            avg_resolution
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    today_str = date.today().strftime("%d-%m-%Y")
    response['Content-Disposition'] = f'attachment; filename=Engineer_Report_{today_str}.xlsx'

    wb.save(response)
    return response




from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Engineer

@login_required
def add_engineer_ajax(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()

        if not name:
            return JsonResponse(
                {"error": "Engineer name is required"},
                status=400
            )

        station = request.user.userprofile.station

        
        if Engineer.objects.filter(
            name__iexact=name,
            station=station
        ).exists():
            return JsonResponse(
                {"error": f"Engineer '{name}' already exists for this station"},
                status=400
            )

        engineer = Engineer.objects.create(
            name=name,
            station=station,
            is_active=True
        )

        return JsonResponse({
            "id": engineer.id,
            "name": engineer.name
        })

    return JsonResponse({"error": "Invalid request"}, status=400)



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import MISForm, MISField, MISSubmission



# CREATE MIS FORM 

@login_required
def create_mis_form(request):
    if not getattr(request.user.userprofile, "is_rhqwr", False):
        messages.error(request, "Permission denied.")
        return redirect("dashboard")

    if request.method == 'POST':
        form_name = request.POST.get('form_name')
        labels = request.POST.getlist('label[]')
        types = request.POST.getlist('type[]')

        form = MISForm.objects.create(name=form_name)

        for label, field_type in zip(labels, types):
            MISField.objects.create(
                form=form,
                label=label,
                field_type=field_type
            )

        messages.success(request, "MIS Form created successfully.")
        return redirect('mis_form_list')    

    return render(request, 'accounts/create_form.html')



# FORM LIST (ACTIVE / ARCHIVED)

@login_required
def mis_form_list(request):
    user = request.user

    
    active_forms = MISForm.objects.filter(
        is_archived=False
    ).order_by("-created_at")

    # Archived forms 
    if user.is_superuser or getattr(user.userprofile, "is_rhqwr", False):
        archived_forms = MISForm.objects.filter(
            is_archived=True
        ).order_by("-archived_at")
    else:
        archived_forms = MISForm.objects.none()

    return render(request, "accounts/form_list.html", {
        "active_forms": active_forms,
        "archived_forms": archived_forms,
    })




# FILL MIS FORM 
@login_required
def mis_form_fill(request, form_id):
    form = get_object_or_404(MISForm, id=form_id)

    # Archived form
    if form.is_archived:
        messages.error(request, "This MIS Form is archived.")
        return redirect("mis_form_list")

    profile = request.user.userprofile
    fields = MISField.objects.filter(form=form)

  
    
    if profile.is_rhqwr:

        from .models import Station
        stations = Station.objects.all()

        # Stations which already submitted this form
        submitted_station_ids = set(
            MISSubmission.objects.filter(form=form)
            .values_list("station_id", flat=True)
        )

        if request.method == "POST":
            station_id = request.POST.get("station")
            if not station_id:
                messages.error(request, "Please select a station.")
                return redirect(request.path)

            station = get_object_or_404(Station, id=station_id)

            #  Duplicate safety
            if MISSubmission.objects.filter(form=form, station=station).exists():
                messages.error(request, "This station has already submitted this form.")
                return redirect("mis_form_list")

            data = {}
            for field in fields:
                data[field.label] = request.POST.get(
                    f"field_{field.id}", ""
                ).strip()

            MISSubmission.objects.create(
                form=form,
                station=station,
                submitted_by=request.user,
                data=data
            )

            messages.success(request, "MIS Form submitted successfully.")
            return redirect("mis_form_list")

        return render(request, "accounts/form_fill.html", {
            "form": form,
            "fields": fields,
            "stations": stations,
            "submitted_station_ids": submitted_station_ids,
            "is_rhqwr": True
        })

  
    # STATION USER LOGIC
  
    station = getattr(profile, "station", None)
    if not station:
        messages.error(request, "Station not assigned.")
        return redirect("mis_form_list")

    submission = MISSubmission.objects.filter(
        form=form,
        station=station
    ).first()

    if submission:
        if submission.status == "editable":
            return redirect("mis_form_edit", submission_id=submission.id)
        messages.warning(request, "You have already submitted this form.")
        return redirect("mis_form_list")

    if request.method == "POST":
        data = {}
        for field in fields:
            data[field.label] = request.POST.get(
                f"field_{field.id}", ""
            ).strip()

        MISSubmission.objects.create(
            form=form,
            station=station,
            submitted_by=request.user,
            data=data
        )

        messages.success(request, "MIS Form submitted successfully.")
        return redirect("mis_form_list")

    return render(request, "accounts/form_fill.html", {
        "form": form,
        "fields": fields,
        "is_rhqwr": False
    })




# EDIT MIS FORM 
@login_required
def mis_form_edit(request, submission_id):
    submission = get_object_or_404(MISSubmission, id=submission_id)

    userprofile = request.user.userprofile
    if submission.station != getattr(userprofile, "station", None):
        messages.error(request, "Unauthorized access.")
        return redirect("mis_form_list")

    if submission.status != "editable":
        messages.error(request, "Edit not allowed.")
        return redirect("mis_form_list")

    fields = MISField.objects.filter(form=submission.form)

    if request.method == "POST":
        updated_data = {}
        for field in fields:
            updated_data[field.label] = request.POST.get(
                f"field_{field.id}", ""
            ).strip()

        submission.data = updated_data
        submission.status = "submitted"
        submission.save()

        messages.success(request, "MIS Form updated successfully.")
        return redirect("mis_form_list")

    field_values = []
    for field in fields:
        field_values.append({
            "id": field.id,
            "label": field.label,
            "field_type": field.field_type,
            "value": submission.data.get(field.label, "")
        })

    return render(request, "accounts/form_fill.html", {
        "form": submission.form,
        "field_values": field_values,
        "is_edit": True
    })





@login_required
def allow_mis_edit(request, submission_id):
    if not getattr(request.user.userprofile, "is_rhqwr", False):
        messages.error(request, "Permission denied.")
        return redirect("mis_submission_list")

    submission = get_object_or_404(MISSubmission, id=submission_id)

    submission.status = "editable"
    submission.edit_allowed_by = request.user
    submission.edit_allowed_at = timezone.now()
    submission.save()

    messages.success(request, "Edit permission granted.")
    return redirect("mis_submission_list")



# MIS SUBMISSION LIST
@login_required
def mis_submission_list(request):
    user = request.user
    profile = user.userprofile

    submissions = MISSubmission.objects.select_related(
        "form", "submitted_by"
    ).filter(form__is_archived=False)

    if not (user.is_superuser or profile.is_rhqwr):
        station = getattr(profile, "station", None)
        submissions = submissions.filter(station=station) if station else submissions.none()

    submissions = submissions.order_by("-submitted_at")

    grouped_submissions = {}
    for sub in submissions:
        grouped_submissions.setdefault(sub.form, []).append(sub)

    return render(request, "accounts/submission_list.html", {
        "grouped_submissions": grouped_submissions
    })



# EXPORT EXCEL  

@login_required
def export_mis_form_excel(request, form_id):
    form = get_object_or_404(MISForm, id=form_id)

    if form.is_archived:
        return HttpResponse("This form is archived.")

    user = request.user
    profile = user.userprofile

    # Role-based data access
    if user.is_superuser or profile.is_rhqwr:
        submissions = MISSubmission.objects.filter(form=form)
    else:
        submissions = MISSubmission.objects.filter(
            form=form,
            station=profile.station
        )

    if not submissions.exists():
        return HttpResponse("No data available.")

    wb = Workbook()
    ws = wb.active
    ws.title = form.name[:30]

    #  Headers 
    field_keys = list(submissions.first().data.keys())
    headers = [
        "Form Name",
        "Station",
        "Submitted By",
        *field_keys,
        "Submitted Date & Time"
    ]
    ws.append(headers)

    # Header Styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="BDD7EE",
            end_color="BDD7EE",
            fill_type="solid"
        )

    #  Data Rows
    for sub in submissions:
        row = [
            form.name,
            str(sub.station),
            sub.submitted_by.username if sub.submitted_by else ""
        ]

        for key in field_keys:
            row.append(sub.data.get(key, ""))

       
        row.append(
            timezone.localtime(sub.submitted_at).strftime('%d-%m-%Y %I:%M %p')
            if sub.submitted_at else ""
        )

        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{form.name}_MIS.xlsx"'
    wb.save(response)
    return response

# ARCHIVE / UNARCHIVE FORM (RHQWR)

@login_required
def archive_mis_form(request, form_id):
    if not getattr(request.user.userprofile, "is_rhqwr", False):
        messages.error(request, "Permission denied.")
        return redirect("mis_form_list")

    form = get_object_or_404(MISForm, id=form_id)
    form.is_archived = True
    form.archived_at = timezone.now()
    form.save()

    messages.success(request, "MIS Form archived.")
    return redirect("mis_form_list")


@login_required
def unarchive_mis_form(request, form_id):
    if not getattr(request.user.userprofile, "is_rhqwr", False):
        messages.error(request, "Permission denied.")
        return redirect("mis_form_list")

    form = get_object_or_404(MISForm, id=form_id)
    form.is_archived = False
    form.archived_at = None
    form.save()

    messages.success(request, "MIS Form unarchived.")
    return redirect("mis_form_list")


# Edit mis submission
@login_required
def edit_mis_submission(request, submission_id):
    submission = get_object_or_404(MISSubmission, id=submission_id)

    user_station = request.user.userprofile.station
    is_rhqwr = request.user.userprofile.is_rhqwr

    # Unauthorized access
    if not is_rhqwr and submission.station != user_station:
        return render(request, 'permission_denied.html')

    # Status check
    if submission.status not in ['submitted', 'editable']:
        return render(request, 'permission_denied.html')

    fields = submission.form.misfield_set.all()

    if request.method == 'POST':
        updated_data = {}
        for field in fields:
            # Station field is not editable
            if field.label.lower() == 'station name':
                updated_data[field.label] = submission.data.get(field.label, '')
            else:
                updated_data[field.label] = request.POST.get(field.label, '').strip()

        submission.data = updated_data
        submission.status = 'editable'
        submission.edit_allowed_by = request.user
        submission.edit_allowed_at = timezone.now()
        submission.save()
        return redirect('mis_submission_list')

    # Prepare data_items for template 
    data_items = []
    for field in fields:
        value = submission.data.get(field.label, '')
        is_editable = field.label.lower() != 'station name'
        data_items.append({
            'label': field.label,
            'value': value,
            'editable': is_editable,
            'type': field.field_type
        })

    context = {
        'submission': submission,
        'data_items': data_items,
    }
    return render(request, 'accounts/edit_mis_submission.html', context)



